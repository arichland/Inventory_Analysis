_author_ = 'arichland'

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pymysql.cursors
import pandas as pd
import numpy as np
import pydict
from datetime import date
import pprint
pp = pprint.PrettyPrinter(indent=1)

# OPENPYXL FIELDS
# Setup workbook connection
filename = "Inventory Analysis.xlsx"
wb = openpyxl.load_workbook(filename=filename)
df_rows = dataframe_to_rows

# Create styles for formatting cells
styles = openpyxl.styles
font = styles.Font
color = styles.Color
align = styles.Alignment
border = styles.Border
side = styles.Side
colors = styles.Color
namedstyle = styles.NamedStyle
header = namedstyle(name="header")
headers = namedstyle(name="cols")
headers.font = font(bold=True)
headers.alignment = align(horizontal="center", vertical="center")
headers.border = border(bottom=side(border_style="thin"))

class sql:
    def query(str):
        local = pydict.localhost.get
        user = local('user')
        password = local('password')
        host = local('host')
        database = local('database')
        data = {}
        headers = []
        con = pymysql.connect(user=user, password=password, host=host, database=database)
        try:
         with con.cursor() as cur:
            # Inventory History Query
            query = pydict.queries.get(str)
            cur.execute(query) # run SQL query
            rows = cur.fetchall() # Assign variable for rows
            cols = cur.description # Assign variable for columns

            for i in range(len(cols)):
                headers.append(cols[i][0])

            for row in rows: # SQL query to dict
                temp_dict = {row[0]: {
                    headers[0]: row[0],
                    headers[1]: row[1],
                    headers[2]: row[2],
                    headers[3]: row[3],
                    headers[4]: row[4],
                    headers[5]: row[5],
                    headers[6]: row[6]}}
                data.update(temp_dict)
        finally:
            con.commit()
            cur.close()
            con.close()
        return data

class analysis:

    def inventory_history():
        print(" Inventory History")
        inven_hist = sql.query("inventory history")
        wb.active = wb.sheetnames.index("Inventory History")
        ws = wb.active

        # Load inven_history_dict to Pandas dataframe
        inven_hist_df = pd.DataFrame.from_dict(inven_hist, orient="index")
        inven_hist_df = pd.pivot_table(inven_hist_df, values='qty', index=['concat', 'sku', 'location'], columns=['dates'], aggfunc=np.sum, fill_value=0)
        num_cols = len(inven_hist_df.columns)

        # Pandas calculations
        inven_hist_df['Average'] = inven_hist_df.iloc[:, 0:num_cols].mean(axis=1) # Avg inventory per SKU, per location
        inven_hist_df['Min'] = inven_hist_df.iloc[:, 0:num_cols].min(axis=1) # Min inventory quantity per SKU, per location
        inven_hist_df['Max'] = inven_hist_df.iloc[:, 0:num_cols].max(axis=1)  # Max inventory quantity per SKU, per location
        inven_hist_df["Total"] = inven_hist_df.iloc[:, 0:num_cols].sum(axis=1) # Sum of inventory per month, per SKU per location
        inven_hist_df["Months w/ Inven"] = inven_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1) # Count of months with inventory quantities greater than 0
        inven_hist_df["Avg Inven when >0"] = (inven_hist_df.iloc[:, 0:num_cols].sum(axis=1))/(inven_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1)) # Average inventory quantity of all months with inventory greater than 0

        # Write dataframe to workbook
        writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
        book = wb
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        inven_hist_df.to_excel(writer, "Inventory History",  startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))
        writer.save()

        # Format rows and columns with openpyxl
        for column in ws["A:C"]:
            for cell in column:
                cell.font = font(bold=None)
                cell.alignment = align(horizontal="left", vertical="center")
                cell.border = border(left=None, right=None, top=None, bottom=None)

        for cell in ws[3]:
            cell.font = headers.font
            cell.alignment = headers.alignment
            cell.border = headers.border

    def current_inven():
        print(" Current Inventory")
        current_inven = sql.query("current inventory")
        # CURRENT INVENTORY
        wb.active = wb.sheetnames.index("Current Inventory")
        ws = wb.active

        # Load to Pandas dataframe
        cur_inven_df = pd.DataFrame.from_dict(current_inven, orient="index")
        cur_inven_df = pd.pivot_table(cur_inven_df, values='qty', index=['concat', 'sku', 'location', 'cogs', 'msrp'], fill_value=0)
        num_cols = len(cur_inven_df.columns)

        writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
        book = wb
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        cur_inven_df.to_excel(writer, "Current Inventory",  startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))
        writer.save()

        # Format rows and columns with openpyxl
        for column in ws["A:E"]:
            for cell in column:
                cell.font = font(bold=None)
                cell.alignment = align(horizontal="left", vertical="center")
                cell.border = border(left=None, right=None, top=None, bottom=None)

        for cell in ws[3]:
            cell.font = headers.font
            cell.alignment = headers.alignment
            cell.border = headers.border
        wb.save(filename=filename)

    def sales_history():
        print(" Sales History")
        sales_hist = sql.query("sales history")

        # SALES HISTORY
        wb.active = wb.sheetnames.index("Sales History")
        ws = wb.active

        # Load to Pandas dataframe
        sales_hist_df = pd.DataFrame.from_dict(sales_hist, orient="index")
        sales_hist_df = pd.pivot_table(sales_hist_df, index=['concat', 'sku', 'location'], columns=['dates'], values='sales', aggfunc=np.sum, fill_value=0)
        num_cols = len(sales_hist_df.columns)

        # Pandas calculations
        sales_hist_df["Total Demand"] = sales_hist_df.iloc[:, 0:num_cols].sum(axis=1) # Sum of demand per month, per SKU per location
        sales_hist_df["Active Months"] = sales_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1) # Count of months with sales

        sales_hist_df['Average'] = sales_hist_df.iloc[:, 0:num_cols].mean(axis=1) # Avg demand per SKU, per location
        sales_hist_df['Min'] = sales_hist_df.iloc[:, 0:num_cols].min(axis=1) # Min sales
        sales_hist_df['Max'] = sales_hist_df.iloc[:, 0:num_cols].max(axis=1)  # Max sales

        sales_hist_df["Months w/ Inven"] = sales_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1) # Count of months with inventory quantities greater than 0
        sales_hist_df["Avg Inven when >0"] = (sales_hist_df.iloc[:, 0:num_cols].sum(axis=1))/(sales_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1)) # Average inventory quantity of all months with inventory greater than 0

        # Write dataframe to workbook
        writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
        book = wb
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        sales_hist_df.to_excel(writer, "Sales History",  startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))
        writer.save()

        # Format rows and columns with openpyxl
        for column in ws["A:C"]:
            for cell in column:
                cell.font = font(bold=None)
                cell.alignment = align(horizontal="left", vertical="center")
                cell.border = border(left=None, right=None, top=None, bottom=None)
        for cell in ws[3]:
            cell.font = headers.font
            cell.alignment = headers.alignment
            cell.border = headers.border

        wb.save(filename=filename)

    def forecasts():
        print(" Forecasts")
        fc = sql.query("forecasts")
        wb.active = wb.sheetnames.index("Forecast History")
        ws = wb.active

        # Load inven_history_dict to Pandas dataframe
        df = pd.DataFrame.from_dict(fc, orient="index")
        df = pd.pivot_table(df, values='forecast', index=['concat', 'sku', 'location'], columns=['dates'], aggfunc=np.sum, fill_value=0)
        num_cols = len(df.columns)

        writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
        book = wb
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, "Forecast History", startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))


        # Format rows and columns with openpyxl
        for column in ws["A:C"]:
            for cell in column:
                cell.font = font(bold=None)
                cell.alignment = align(horizontal="left", vertical="center")
                cell.border = border(left=None, right=None, top=None, bottom=None)

        for cell in ws[3]:
            cell.font = headers.font
            cell.alignment = headers.alignment
            cell.border = headers.border
        writer.save()

    def materials():
        print(" Materials")
        materials = sql.query("materials")

        # Set active sheet
        wb.active = wb.sheetnames.index("Material Data")
        ws = wb.active

        df = pd.DataFrame.from_dict(materials, orient="index")
        df  = df.set_index('material')
        df = pd.DataFrame(df, columns=['description', 'unit_cost', 'currency', 'unit_price', 'category', 'subcategory', 'org_code'])

        writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
        book = wb
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, "Material Data", startrow=2, startcol=0, engine="openpyxl",
                        freeze_panes=(3, 1))
        for column in ws["A:B"]:
            for cell in column:
                cell.font = font(bold=None)
                cell.alignment = align(horizontal="left", vertical="center")
                cell.border = border(left=None, right=None, top=None, bottom=None)

        for cell in ws[3]:
            cell.font = headers.font
            cell.alignment = headers.alignment
            cell.border = headers.border
        writer.save()

    def skus():
       print(" SKUs")
       sku = sql.query("skus")

        # Set active sheet
       wb.active = wb.sheetnames.index("SKU Data")
       ws = wb.active

       df = pd.DataFrame.from_dict(sku, orient="index")
       df = df.set_index('concat')
       df = pd.DataFrame(df, columns=['sku_id', 'location', 'make_buy', 'name', 'description'])

       writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
       book = wb
       writer.book = book
       writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
       df.to_excel(writer, "SKU Data", startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))
       for column in ws["A:B"]:
           for cell in column:
               cell.font = font(bold=None)
               cell.alignment = align(horizontal="left", vertical="center")
               cell.border = border(left=None, right=None, top=None, bottom=None)

       for cell in ws[3]:
           cell.font = headers.font
           cell.alignment = headers.alignment
           cell.border = headers.border
       writer.save()


def functions():
   #analysis.inventory_history()
   #analysis.current_inven()
   analysis.sales_history()
   analysis.forecasts()
   #analysis.materials()
   #analysis.skus()
functions()
