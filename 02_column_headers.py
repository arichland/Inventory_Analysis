_author_ = 'arichland'

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors



# Call workbook
filename = "Inventory Analysis.xlsx"
wb = load_workbook(filename=filename)

# Assign sheets to variables
params = wb["Parameters"]
analysis = wb["Analysis"]
inven_hist = wb["Inventory History"]
cur_inven = wb["Current Inventory"]
sales_hist = wb["Sales History"]
forecast_hist = wb["Forecast History"]
inven_turns = wb["Inventory Turns"]
segmen = wb["Segmentation"]
segmen_calc = wb["Segmentation Calculation"]
material = wb["Material Data"]
skus = wb["SKU Data"]
locale = wb["Location Data"]
conversion = wb["Conversion"]



# Format Analysis
analysis.freeze_panes = "C4"
analysis['B3'] = 'Concat'
analysis['C3'] = 'Material'
analysis['D3'] = 'Location'
analysis['E3'] = 'Material Description'
analysis['F3'] = 'Location Desctiption'
analysis['G3'] = 'Include?'
analysis['H3'] = 'Purchased/Produced?'
analysis['I3'] = 'Category'
analysis['J3'] = 'Sub Category'
analysis['K3'] = 'Abc Class'
analysis['L3'] = 'Inven Turns'
analysis['M3'] = 'Avg Days To Turn Inven'
analysis['N3'] = 'Velocity Category'
analysis['O3'] = 'Stock In Months'
analysis['P3'] = 'Stock Category'
analysis['Q3'] = 'Abc Class Category'
analysis['R3'] = 'Unit Price'
analysis['S3'] = 'Unit Cost'
analysis['T3'] = 'Current Inventory'
analysis['U3'] = 'Inventory Value'
analysis['V3'] = 'Average Inventory'
analysis['W3'] = 'Average Value'
analysis['X3'] = 'Annual Volume'
analysis['Y3'] = 'Demand $'
analysis['Z3'] = 'Annual Margin'
analysis['AA3'] = 'Months Active'
analysis['AB3'] = 'Demand Variability'
analysis['AC3'] = 'Fcst Accuracy'
analysis['AD3'] = 'SS - Demand Variance'
analysis['AE3'] = 'SS - Forecast Error'
analysis['AF3'] = 'Suggested SS Method'
analysis['AG3'] = 'Cycle Stock'
analysis['AH3'] = 'Min Inventory'
analysis['AI3'] = 'Max Inventory'
analysis['AJ3'] = 'Avg Inventory'
analysis['AK3'] = 'Avg Inventory Value'
analysis['AL3'] = 'Diff vs Current'



def analysis_header():
    worksheet_names = wb.sheetnames
    sheet_index = worksheet_names.index("Analysis")
    wb.active = sheet_index
    sheet = wb.active
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.align = Alignment(horizontal="center", vertical="center")
    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

# Format Inventory History
inven_hist.freeze_panes = "C4"
inven_hist['B3'] = 'Concat'
inven_hist['C3'] = 'Material'
inven_hist['D3'] = 'Location'
inven_hist['E3'] = 'Month 01'
inven_hist['F3'] = 'Month 02'
inven_hist['G3'] = 'Month 03'
inven_hist['H3'] = 'Month 04'
inven_hist['I3'] = 'Month 05'
inven_hist['J3'] = 'Month 06'
inven_hist['K3'] = 'Month 07'
inven_hist['L3'] = 'Month 08'
inven_hist['M3'] = 'Month 09'
inven_hist['N3'] = 'Month 10'
inven_hist['O3'] = 'Month 11'
inven_hist['P3'] = 'Month 12'
inven_hist['Q3'] = 'Month 13'
inven_hist['R3'] = 'Month 14'
inven_hist['S3'] = 'Average'
inven_hist['T3'] = 'Min'
inven_hist['U3'] = 'Max'
inven_hist['V3'] = 'Total'
inven_hist['W3'] = 'Months with inventory'
inven_hist['X3'] = 'Avg Inventory when >0'
inven_hist['Y3'] = 'Min when >0'


def inven_hist_header():
    worksheet_names = wb.sheetnames
    sheet_index = worksheet_names.index("Inventory History")
    wb.active = sheet_index
    sheet = wb.active
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.align = Alignment(horizontal="center", vertical="center")
    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

# Format Current Inventory
cur_inven.freeze_panes = "C4"
cur_inven['B3'] = 'Concat'
cur_inven['C3'] = 'Material'
cur_inven['D3'] = 'Location'
cur_inven['E3'] = 'Current Inventory'
cur_inven['F3'] = 'Inventory value'
cur_inven['G3'] = 'COGS'
cur_inven['H3'] = 'Price'


# Format Sales History
sales_hist.freeze_panes = "C4"
sales_hist['B3'] = 'Concat'
sales_hist['C3'] = 'Material'
sales_hist['D3'] = 'Location'
sales_hist['E3'] = 'Month 01'
sales_hist['F3'] = 'Month 02'
sales_hist['G3'] = 'Month 03'
sales_hist['H3'] = 'Month 04'
sales_hist['I3'] = 'Month 05'
sales_hist['J3'] = 'Month 06'
sales_hist['K3'] = 'Month 07'
sales_hist['L3'] = 'Month 08'
sales_hist['M3'] = 'Month 09'
sales_hist['N3'] = 'Month 10'
sales_hist['O3'] = 'Month 11'
sales_hist['P3'] = 'Month 12'
sales_hist['Q3'] = 'Total Demand'
sales_hist['R3'] = 'Active Months'
sales_hist['S3'] = 'Avg Demand'
sales_hist['T3'] = 'Min'
sales_hist['U3'] = 'Max'
sales_hist['V3'] = 'AVG >0'
sales_hist['W3'] = 'MIN >0'
sales_hist['X3'] = 'Lt'
sales_hist['Y3'] = 'Annual'
sales_hist['Z3'] = 'Q1'
sales_hist['AA3'] = 'Q2'
sales_hist['AB3'] = 'Q3'
sales_hist['AC3'] = 'Q4'
sales_hist['AD3'] = 'Annual'
sales_hist['AE3'] = 'Q1'
sales_hist['AF3'] = 'Q2'
sales_hist['AG3'] = 'Q3'
sales_hist['AH3'] = 'Q4'
sales_hist['AI3'] = 'Annual'
sales_hist['AJ3'] = 'Q1'
sales_hist['AK3'] = 'Q2'
sales_hist['AL3'] = 'Q3'


# Format Forecast History
forecast_hist.freeze_panes = "C4"
forecast_hist['B3'] = 'Concat'
forecast_hist['C3'] = 'Material'
forecast_hist['D3'] = 'Location'
forecast_hist['E3'] = 'Month 01'
forecast_hist['F3'] = 'Month 02'
forecast_hist['G3'] = 'Month 03'
forecast_hist['H3'] = 'Month 04'
forecast_hist['I3'] = 'Month 05'
forecast_hist['J3'] = 'Month 06'
forecast_hist['K3'] = 'Month 07'
forecast_hist['L3'] = 'Month 08'
forecast_hist['M3'] = 'Month 09'
forecast_hist['N3'] = 'Month 10'
forecast_hist['O3'] = 'Month 11'
forecast_hist['P3'] = 'Month 12'
forecast_hist['Q3'] = 'Month 01'
forecast_hist['R3'] = 'Month 02'
forecast_hist['S3'] = 'Month 03'
forecast_hist['T3'] = 'Month 04'
forecast_hist['U3'] = 'Month 05'
forecast_hist['V3'] = 'Month 06'
forecast_hist['W3'] = 'Month 07'
forecast_hist['X3'] = 'Month 08'
forecast_hist['Y3'] = 'Month 09'
forecast_hist['Z3'] = 'Month 10'
forecast_hist['AA3'] = 'Month 11'
forecast_hist['AB3'] = 'Month 12'
forecast_hist['AC3'] = 'Month 01'
forecast_hist['AD3'] = 'Month 02'
forecast_hist['AE3'] = 'Month 03'
forecast_hist['AF3'] = 'Month 04'
forecast_hist['AG3'] = 'Month 05'
forecast_hist['AH3'] = 'Month 06'
forecast_hist['AI3'] = 'Month 07'
forecast_hist['AJ3'] = 'Month 08'
forecast_hist['AK3'] = 'Month 09'
forecast_hist['AL3'] = 'Month 10'
forecast_hist['AM3'] = 'Month 11'
forecast_hist['AN3'] = 'Month 12'
forecast_hist['AO3'] = 'Total Bias'
forecast_hist['AP3'] = 'Month 01'
forecast_hist['AQ3'] = 'Month 02'
forecast_hist['AR3'] = 'Month 03'
forecast_hist['AS3'] = 'Month 04'
forecast_hist['AT3'] = 'Month 05'
forecast_hist['AU3'] = 'Month 06'
forecast_hist['AV3'] = 'Month 07'
forecast_hist['AW3'] = 'Month 08'
forecast_hist['AX3'] = 'Month 09'
forecast_hist['AY3'] = 'Month 10'
forecast_hist['AZ3'] = 'Month 11'
forecast_hist['BA3'] = 'Month 12'
forecast_hist['BB3'] = 'RMSE'
forecast_hist['BC3'] = 'RMSE ACTIVE'
forecast_hist['BD3'] = 'Month 01'
forecast_hist['BE3'] = 'Month 02'
forecast_hist['BF3'] = 'Month 03'
forecast_hist['BG3'] = 'Month 04'
forecast_hist['BH3'] = 'Month 05'
forecast_hist['BI3'] = 'Month 06'
forecast_hist['BJ3'] = 'Month 07'
forecast_hist['BK3'] = 'Month 08'
forecast_hist['BL3'] = 'Month 09'
forecast_hist['BM3'] = 'Month 10'
forecast_hist['BN3'] = 'Month 11'
forecast_hist['BO3'] = 'Month 12'
forecast_hist['BP3'] = 'SUM OF ERROR %'
forecast_hist['BQ3'] = 'AVG MAPE'
forecast_hist['BR3'] = 'Avg Bias'
forecast_hist['BS3'] = 'AVG MAPE'
forecast_hist['BT3'] = 'Avg Bias'
forecast_hist['BU3'] = 'MAD'
forecast_hist['BV3'] = 'Tracking Signal'
forecast_hist['BW3'] = 'TS'
forecast_hist['BX3'] = 'TY Accuracy'
forecast_hist['BY3'] = 'Active Accuracy'
forecast_hist['BZ3'] = 'SS ANNUAL ERROR'
forecast_hist['CA3'] = 'SS ACTIVE ERROR'


#Format Inventory Turns
inven_turns['B3'] = 'Concat'
inven_turns['C3'] = 'Material'
inven_turns['D3'] = 'Location'
inven_turns['E3'] = 'Qty On Hand'
inven_turns['F3'] = 'On Hand Value'
inven_turns['G3'] = 'COGS'
inven_turns['H3'] = 'Avg Inven $'
inven_turns['I3'] = 'Inven Turns'
inven_turns['J3'] = 'Days to Turn Inven'
inven_turns['K3'] = 'Velocity Category'
inven_turns['L3'] = 'On Hand Inventory'
inven_turns['M3'] = 'Stock in Months'
inven_turns['N3'] = 'Stock Category'


# Segmentation
segmen.freeze_panes = "C4"
segmen['B3'] = 'Concat'
segmen['C3'] = 'Sku'
segmen['D3'] = 'Loc'
segmen['E3'] = 'Segmentation'
segmen['F3'] = 'Category Segmentation'


# Segmentation Calculation
segmen_calc.freeze_panes = "C4"
segmen_calc['B3'] = 'Concat'
segmen_calc['C3'] = 'Material'
segmen_calc['D3'] = 'Location'
segmen_calc['E3'] = 'Category'
segmen_calc['F3'] = 'Sub Category'
segmen_calc['G3'] = 'Annual Volume'
segmen_calc['H3'] = 'Demand $'
segmen_calc['I3'] = 'Annual Margin'
segmen_calc['J3'] = 'Sku %'
segmen_calc['K3'] = 'Cumulative'

# Material Data
material.freeze_panes = "C4"
material['B3'] = 'Material'
material['C3'] = 'Description'
material['D3'] = 'Unit Cost'
material['E3'] = 'Currency'
material['F3'] = 'Unit Price'
material['G3'] = 'Category'
material['H3'] = 'Subcategory'
material['I3'] = 'Status'
material['J3'] = 'Unit Price'
material['K3'] = 'Units/Case'
material['L3'] = 'Case/Pallet'

# SKU Data
skus.freeze_panes = "C4"
skus['B3'] = 'ConCat'
skus['C3'] = 'Material'
skus['D3'] = 'Locale'
skus['E3'] = 'Purchasing/Producing?'

# Locales
locale.freeze_panes = "C4"
locale['B3'] = 'Location'
locale['C3'] = 'Description'

index_params = 0
index_analysis = 1
index_inven_hist = 2
index_cur_inven = 3
index_sales_hist = 4
index_forecast_hist = 5
index_inven_turns = 6
index_segmen = 7
index_segmen_calc = 8
index_material = 9
index_skus = 10
index_locale = 11
index_conversion = 12
wb.save(filename=filename)