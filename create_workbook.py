_author_ = 'arichland'

import openpyxl
class workbook:
    wb = openpyxl.Workbook()
    ws = openpyxl.Workbook.active
    filename = "Inventory Analysis.xlsx"
    styles = openpyxl.styles
    font = styles.Font
    color = styles.Color
    align = styles.Alignment
    border = styles.Border
    side = styles.Side
    colors = styles.Color
    namedstyle = styles.NamedStyle
    header = namedstyle(name="header")
    header.font = font(bold=True)
    header.alignment = align(horizontal="left", vertical="center")

    cols = namedstyle(name="cols")
    cols.font = font(bold=True)
    cols.alignment = align(horizontal="center", vertical="center")
    cols.border = border(bottom=side(border_style="thin"))
    sheets = {
        "Parameters": 0,
        "Analysis": 1,
        "Inventory History": 2,
        "Current Inventory": 3,
        "Sales History": 4,
        "Forecast History": 5,
        "Inventory Turns": 6,
        "Segmentation": 7,
        "Segmentation Calculation": 8,
        "Material Data": 9,
        "SKU Data": 10,
        "Location Data": 11,
        "Conversion": 12}

    def create_sheets():
        wb = workbook.wb
        sheets = workbook.sheets
        filename = workbook.filename
        # Create, name, and put sheets in order
        sheets = workbook.sheets
        for k,v in sheets.items():
            wb.create_sheet(k, v)
        wb.save(filename=filename)

    def params():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Parameters")
        params = wb.active
        header = workbook.header
        cols = workbook.cols
        filename = workbook.filename

        params['A1'] = 'Instructions:'
        params['A1'].style = header

        # Standard Deviation:
        params['A11'] = '12 Month StdDev or Active Month'
        params['A12'] = 'Lead Times'
        params['A13'] = 'Segmentation'
        params['B11'] = '12 Month'
        params['B12'] = 'Universal'
        params['B13'] = 'Overall'

        # CV Variability
        params['D11'] = 'CV Variability Scale'
        params['D11'].style = header
        params['D12'] = 'Very Low'
        params['D13'] = 'Low'
        params['E12'] = .3
        params['E13'] = 1

        # Lead Times
        params['G11'] = 'Lead Times'
        params['G11'].style = header
        params['G12'] = 'Production'
        params['H12'] = 7
        params['G13'] = 'Purchasing'
        params['H13'] = 7
        params['G14'] = 'Transit'
        params['H14'] = 21

        # Tracking Signal
        params['A17'] = 'Tracking Signal'
        params['A17'].style = header
        params['A18'] = 'Extremely High'
        params['A19'] = 'Ver High'
        params['A20'] = 'High'
        params['A21'] = 'Ok +'
        params['A22'] = 'Ok -'
        params['A23'] = 'High -'
        params['A24'] = 'Very High -'
        params['A24'] = 'Extremely High'

        params['B18'] = 6
        params['B19'] = 5
        params['B20'] = 4
        params['B21'] = 3
        params['B22'] = -3
        params['B23'] = -4
        params['B24'] = -5
        params['B24'] = -6

        # Forecast Error
        params['D17'] = 'Forecast Error Scale'
        params['D17'].style = header
        params['D18'] = 'Highly Accurate'
        params['D19'] = 'Accurate'
        params['D20'] = 'Average'
        params['D21'] = 'Poor'
        params['D22'] = 'Very Poor'
        params['D23'] = 'Extremely Poor'

        params['E18'] = .15
        params['E19'] = .25
        params['E20'] = .35
        params['E21'] = .55
        params['E22'] = .75
        params['E23'] = 1

        params['E18'].style = 'Percent'
        params['E19'].style = 'Percent'
        params['E20'].style = 'Percent'
        params['E21'].style = 'Percent'
        params['E22'].style = 'Percent'
        params['E23'].style = 'Percent'

        # Inventory Turn Categories
        params['G17'] = 'Inventory Turn Category'
        params['G18'] = 'Category'
        params['G19'] = 'Good'
        params['G20'] = 'Slow Moving'
        params['G21'] = 'Obsolete'
        params['G17'].style = header
        params['G18'].style = cols

        params['H18'] = 'Min'
        params['H19'] = 0
        params['H20'] = 61
        params['H21'] = 121
        params['H18'].style = cols

        params['I18'] = 'Max'
        params['I19'] = 60
        params['I20'] = 120
        params['I18'].style = cols

        # Stock Categories
        params['A27'] = 'Stock Categories'
        params['A28'] = 'Category'
        params['A29'] = 'Understock'
        params['A30'] = 'Good'
        params['A31'] = 'Overstock'
        params['A27'].style = header
        params['A28'].style = cols

        params['B28'] = 'Min'
        params['B29'] = 0
        params['B30'] = 1.5
        params['B31'] = 3
        params['B28'].style = cols

        params['C28'] = 'Max'
        params['C29'] = 1.5
        params['C30'] = 3
        params['C28'].style = cols

        # Months of Stock
        params['E27'] = 'Months of Stock'
        params['E28'] = 'Category'
        params['E29'] = 'Good'
        params['E30'] = 'Slow Moving'
        params['E31'] = 'Obsolete'
        params['E27'].style = header
        params['E28'].style = cols

        params['F28'] = 'Min'
        params['F29'] = 0
        params['F30'] = 1.5
        params['F31'] = 3
        params['F28'].style = cols

        params['G28'] = 'Max'
        params['G29'] = 1.5
        params['G30'] = 3
        params['G28'].style = cols

        # Weighted Service Level
        params['A33'] = 'Weighted Service Level Count'
        params['B33'] = 'Weighted Service Volume'
        params['C33'] = 'Count'
        params['D33'] = 'Percentage of SKUs'
        params['E33'] = 'Weighted Service Level Count'

        # Segmentation: Categories
        params['A41'] = 'Segmentation Service Levels'
        params['A42'] = 'Category'
        params['A43'] = 'AAA'
        params['A44'] = 'A'
        params['A45'] = 'B'
        params['A46'] = 'C'
        params['A47'] = 'E'
        params['A41'].style = header
        params['A42'].style = cols

        # Segmentation: Service Levels
        params['B42'] = 'Service Level'
        params['B43'] = .995
        params['B44'] = .98
        params['B45'] = .965
        params['B46'] = .95
        params['B47'] = 0
        params['B42'].style = cols
        params['B43'].style = 'Percent'
        params['B44'].style = 'Percent'
        params['B45'].style = 'Percent'
        params['B46'].style = 'Percent'
        params['B47'].style = 'Percent'

        # Segmentation: Pareto Level
        params['C42'] = 'Pareto'
        params['C43'] = .25
        params['C44'] = .35
        params['C45'] = .9
        params['C46'] = 1
        params['C42'].style = cols
        params['C43'].style = 'Percent'
        params['C44'].style = 'Percent'
        params['C45'].style = 'Percent'
        params['C46'].style = 'Percent'

        # Segmentation: Cumulative Percentage
        params['D42'] = 'Cumulative'
        params['D43'] = .25
        params['D44'] = .6
        params['D45'] = .9
        params['D46'] = 1
        params['D42'].style = cols
        params['D43'].style = 'Percent'
        params['D44'].style = 'Percent'
        params['D45'].style = 'Percent'
        params['D46'].style = 'Percent'

        # Segmentation: Count of SKUs
        params['E42'] = 'Count of SKUs'
        params['E43'] = 0
        params['E44'] = 0
        params['E45'] = 0
        params['E46'] = 0
        params['E42'].style = cols

        # Segmentation: Percentage of SKUs
        params['F42'] = '% of SKUs'
        params['F43'] = 0
        params['F44'] = 0
        params['F45'] = 0
        params['F46'] = 0
        params['F42'].style = cols
        params['F43'].style = 'Percent'
        params['F44'].style = 'Percent'
        params['F45'].style = 'Percent'
        params['F46'].style = 'Percent'

        # Exchange Rates
        params['A49'] = 'Exchange Rates to USD'
        params['A50'] = 'Currency'
        params['B50'] = 'Country'
        params['C50'] = 'Rate'
        params['A49'].style = header
        params['A50'].style = cols
        params['B50'].style = cols
        params['C50'].style = cols

        wb.save(filename=filename)

    def analysis():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Analysis")
        ws = wb.active
        print("Analysis")

    def inven_history():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Inventory History")
        ws = wb.active
        print("Inventory History")

    def current_inven():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Current Inventory")
        ws = wb.active
        print("Current Inventory")

    def sales_history():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Sales History")
        ws = wb.active
        print("Sales History")

    def forecast_history():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Forecast History")
        ws = wb.active
        print("Forecast History")

    def inven_turns():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Inventory Turns")
        ws = wb.active
        print("Inventory Turns")

    def segmen():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Segmentation")
        ws = wb.active
        print("Segmentation")

    def segmen_calc():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Segmentation Calculation")
        ws = wb.active
        print("Segmentation Calculation")

    def mat_data():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Material Data")
        ws = wb.active
        print("Material Data")

    def sku_data():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("SKU Data")
        ws = wb.active
        print("SKU Data")

    def location():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Location Data")
        ws = wb.active
        print("Location Data")

    def conversion():
        sheets = workbook.sheets
        wb = workbook.wb
        wb.active = workbook.sheets.get("Conversion")
        ws = wb.active
        print("Conversion")

    def create():
        workbook.create_sheets()
        workbook.params()
        workbook.analysis()
        workbook.inven_history()
        workbook.current_inven()
        workbook.sales_history()
        workbook.forecast_history()
        workbook.inven_turns()
        workbook.segmen()
        workbook.segmen_calc()
        workbook.mat_data()
        workbook.sku_data()
        workbook.location()
        workbook.conversion()
workbook.create()