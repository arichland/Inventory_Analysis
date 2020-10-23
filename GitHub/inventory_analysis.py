_author_ = 'arichland'

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pymysql.cursors
from pydict import sql_dict
import pandas as pd

# SQL Fields
user = sql_dict.get('user')
password = sql_dict.get('password')
host = sql_dict.get('host')
database = sql_dict.get('database')
charset = sql_dict.get('charset')
con = pymysql.connect(user=user,
                       password=password,
                       host=host,
                       database=database,
                       charset=charset)

# Excel Fields
filename = "Inventory Analysis.xlsx"
wb = load_workbook(filename=filename)
wb.active = wb.sheetnames.index("Analysis")
ws = wb.active

try:
    with con.cursor() as cur:
        sum = 'SELECT Concat, Material, Location, Mat_Descr, Location_Descr FROM tbl_Analysis;'
        cur.execute(sum)
        cols = cur.description
        col_name = [col[0] for col in cols]
        data = [dict(zip(col_name, row))
                for row in cur.fetchall()]
        df = pd.DataFrame(data)
        df["Sum"] = df.sum(axis=1)
        rows = dataframe_to_rows(df, index=True, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=2+r_idx, column=c_idx, value=value)
        ws.delete_cols(1)
        ws.delete_rows(4)
        ws.freeze_panes = "B4"
        print("Dataset:")
        print(df)
finally:
    con.commit()
    cur.close()
    con.close()
wb.save(filename=filename)