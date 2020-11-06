_author_ = 'arichland'

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pymysql.cursors
import pandas as pd
import numpy as np
from pydict import sql_dict
import pprint
pp = pprint.PrettyPrinter(indent=1)

# SQL Fields
user = sql_dict.get('user')
password = sql_dict.get('password')
host = sql_dict.get('host')
database = sql_dict.get('database')
charset = sql_dict.get('charset')
con = pymysql.connect(user=user, password=password, host=host, database=database, charset=charset)

# OPENPYXL FIELDS
# Setup workbook connection
filename = "Inventory Analysis.xlsx"
wb = openpyxl.load_workbook(filename=filename)
df_rows = dataframe_to_rows

# Create styles for formatting cells
styles = openpyxl.styles
font = styles.Font
color = styles.Color
align = styles.Alignment
border = styles.Border
side = styles.Side
colors = styles.Color
namedstyle = styles.NamedStyle
header = namedstyle(name="header")
headers = namedstyle(name="cols")
headers.font = font(bold=True)
headers.alignment = align(horizontal="center", vertical="center")
headers.border = border(bottom=side(border_style="thin"))

# Dictionaries to store SQL data for Pandas' dataframes
cols = []
inven_history_dict = {}
current_inven_dict = {}
sales_history_dict = {}

# SQL QUERIES
try:
    with con.cursor() as cur:
        # Inventory History Query
        qry_inven_history = "SELECT id, concat(sku, location) as concat, sku, location, abc_class, dates, qty FROM pyInven_Report order by id ASC;"
        cur.execute(qry_inven_history) # run SQL query
        rows = cur.fetchall() # Assign variable for rows
        desc = cur.description # Assign variable for columns
        qry_cols = desc[0][0], desc[1][0], desc[2][0], desc[3][0], desc[4][0], desc[5][0], desc[6][0]
        cols.append(qry_cols)

        for row in rows: # SQL query to dict
            inven_history_data = {row[0]: {cols[0][0]: row[0],
                              cols[0][1]: row[1],
                              cols[0][2]: row[2],
                              cols[0][3]: row[3],
                              cols[0][4]: row[4],
                              cols[0][5]: row[5],
                              cols[0][6]: row[6]}}
            inven_history_dict.update(inven_history_data) # Save qry_inven_history to inven_history_dict
        cols.clear() # Clear list of columns from qry_inven_history to create dict for next query


        # Current Inventory Query
        qry_current_inven = "SELECT id, concat(sku, location) as concat, sku, location, cogs, msrp, qty FROM pyCurrent_Inven order by id ASC;"
        cur.execute(qry_current_inven)  # run SQL query
        rows = cur.fetchall()  # Assign variable for rows
        desc = cur.description  # Assign variable for columns
        qry_cols = desc[0][0], desc[1][0], desc[2][0], desc[3][0], desc[4][0], desc[5][0], desc[6][0]
        cols.append(qry_cols)

        for row in rows: # SQL query to dict
            current_inven_data = {row[0]: {cols[0][0]: row[0],
                              cols[0][1]: row[1],
                              cols[0][2]: row[2],
                              cols[0][3]: row[3],
                              cols[0][4]: row[4],
                              cols[0][5]: row[5],
                              cols[0][6]: row[6]}}
            current_inven_dict.update(current_inven_data)
            cols.clear()

        # Sales History Query
        qry_sales_hist = "SELECT id, concat(sku, location) as concat, sku, location, dates, sales FROM pySales order by id ASC;"
        cur.execute(qry_current_inven)  # run SQL query
        rows = cur.fetchall()  # Assign variable for rows
        desc = cur.description  # Assign variable for columns
        qry_cols = desc[0][0], desc[1][0], desc[2][0], desc[3][0], desc[4][0], desc[5][0]
        cols.append(qry_cols)

        for row in rows:  # SQL query to dict
            sales_hist_data = {row[0]: {cols[0][0]: row[0],
                                        cols[0][1]: row[1],
                                        cols[0][2]: row[2],
                                        cols[0][3]: row[3],
                                        cols[0][4]: row[4],
                                        cols[0][5]: row[5]}}
            sales_history_dict.update(sales_hist_data)
            cols.clear()
finally:
        con.commit()
        cur.close()
        con.close()


# INVENTORY HISTORY
# Set Inventory History sheet as active sheet
wb.active = wb.sheetnames.index("Inventory History")
ws = wb.active

# Load inven_history_dict to Pandas dataframe
inven_hist_df = pd.DataFrame.from_dict(inven_history_dict, orient="index")
inven_hist_df = pd.pivot_table(inven_hist_df, values='qty', index=['concat', 'sku', 'location'], columns=['dates'], aggfunc=np.sum, fill_value=0)
num_cols = len(inven_hist_df.columns)

# Pandas calculations
inven_hist_df['Average'] = inven_hist_df.iloc[:, 0:num_cols].mean(axis=1) # Avg inventory per SKU, per location
inven_hist_df['Min'] = inven_hist_df.iloc[:, 0:num_cols].min(axis=1) # Min inventory quantity per SKU, per location
inven_hist_df['Max'] = inven_hist_df.iloc[:, 0:num_cols].max(axis=1)  # Max inventory quantity per SKU, per location
inven_hist_df["Total"] = inven_hist_df.iloc[:, 0:num_cols].sum(axis=1) # Sum of inventory per month, per SKU per location
inven_hist_df["Months w/ Inven"] = inven_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1) # Count of months with inventory quantities greater than 0
inven_hist_df["Avg Inven when >0"] = (inven_hist_df.iloc[:, 0:num_cols].sum(axis=1))/(inven_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1)) # Average inventory quantity of all months with inventory greater than 0

# Write dataframe to workbook
writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
book = wb
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
inven_hist_df.to_excel(writer, "Inventory History",  startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))
writer.save()

# Format rows and columns with openpyxl
for column in ws["A:C"]:
   for cell in column:
       cell.font = font(bold=None)
       cell.alignment = align(horizontal="left", vertical="center")
       cell.border = border(left=None, right=None, top=None, bottom=None)
for cell in ws[3]:
    cell.font = headers.font
    cell.alignment = headers.alignment
    cell.border = headers.border


# CURRENT INVENTORY
wb.active = wb.sheetnames.index("Current Inventory")
ws = wb.active

# Load to Pandas dataframe
cur_inven_df = pd.DataFrame.from_dict(current_inven_dict, orient="index")
cur_inven_df = pd.pivot_table(cur_inven_df, values='qty', index=['concat', 'sku', 'location', 'cogs', 'msrp'], fill_value=0)
num_cols = len(cur_inven_df.columns)

writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
book = wb
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
cur_inven_df.to_excel(writer, "Current Inventory",  startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))
writer.save()

# Format rows and columns with openpyxl
for column in ws["A:E"]:
   for cell in column:
       cell.font = font(bold=None)
       cell.alignment = align(horizontal="left", vertical="center")
       cell.border = border(left=None, right=None, top=None, bottom=None)
for cell in ws[3]:
    cell.font = headers.font
    cell.alignment = headers.alignment
    cell.border = headers.border


# SALES HISTORY
wb.active = wb.sheetnames.index("Sales History")
ws = wb.active

# Load to Pandas dataframe
sales_hist_df = pd.DataFrame.from_dict(sales_history_dict, orient="index")
sales_hist_df = pd.pivot_table(sales_history_dict, values='sales', index=['concat', 'sku', 'location'], columns=['dates'], aggfunc=np.sum, fill_value=0)
num_cols = len(sales_hist_df.columns)

# Pandas calculations
sales_hist_df["Total Demand"] = sales_hist_df.iloc[:, 0:num_cols].sum(axis=1) # Sum of demand per month, per SKU per location
sales_hist_df["Active Months"] = sales_hist_df.iloc[:, 0:num_cols].astype(bool).sum(axis=1) # Count of months with sales

# Write dataframe to workbook
writer = pd.ExcelWriter('Inventory Analysis.xlsx', engine='openpyxl')
book = wb
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
inven_hist_df.to_excel(writer, "Sales History",  startrow=2, startcol=0, engine="openpyxl", freeze_panes=(3, 1))
writer.save()

# Format rows and columns with openpyxl
for column in ws["A:C"]:
   for cell in column:
       cell.font = font(bold=None)
       cell.alignment = align(horizontal="left", vertical="center")
       cell.border = border(left=None, right=None, top=None, bottom=None)
for cell in ws[3]:
    cell.font = headers.font
    cell.alignment = headers.alignment
    cell.border = headers.border


wb.save(filename=filename)
print(sales_hist_df)