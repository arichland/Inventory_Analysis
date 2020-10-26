_author_ = 'arichland'

from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors


filename = "Inventory Analysis.xlsx"
wb = load_workbook(filename=filename)
ws = wb["Analysis"]
data = [1, 1]


tbl = Table(displayName="Analysis", ref="A3:Ak5")

style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tbl.tableStyleInfo = style
ws.add_table(tbl)

for row in data:
    ws.append(row)











range = workbook.defined_name.DefinedName
create = wb.defined_names.append


# Create names ranges for each sheet
# Analysis
analysis_concat = range('Analysis_Concat', attr_text='Analysis!$A:$A')
analysis_material = range('Analysis_Material', attr_text='Analysis!$B:$B')
analysis_location = range('Analysis_Location', attr_text='Analysis!$C:$C')
analysis_mat_descr = range('Analysis_Mat_Descr', attr_text='Analysis!$D:$D')
analysis_location_descr = range('Analysis_Location_Descr', attr_text='Analysis!$E:$E')
analysis_include = range('Analysis_Include', attr_text='Analysis!$F:$F')
analysis_purchased = range('Analysis_Purchased', attr_text='Analysis!$G:$G')
analysis_category = range('Analysis_Category', attr_text='Analysis!$H:$H')
analysis_subcategory = range('Analysis_SubCategory', attr_text='Analysis!$I:$I')
analysis_abc_class = range('Analysis_Abc_Class', attr_text='Analysis!$J:$J')
analysis_inven_turns = range('Analysis_Inven_Turns', attr_text='Analysis!$K:$K')
analysis_avg_days = range('Analysis_Avg_Days', attr_text='Analysis!$L:$L')
analysis_velocity_category = range('Analysis_Velocity_Category', attr_text='Analysis!$M:$M')
analysis_stock_in_months = range('Analysis_Stock_In_Months', attr_text='Analysis!$N:$N')
analysis_stock_category = range('Analysis_Stock_Category', attr_text='Analysis!$O:$O')
analysis_abc_class_cat = range('Analysis_Abc_Class_Cat', attr_text='Analysis!$P:$P')
analysis_unit_price = range('Analysis_Unit_Price', attr_text='Analysis!$Q:$Q')
analysis_unit_cost = range('Analysis_Unit_Cost', attr_text='Analysis!$R:$R')
analysis_current_inven = range('Analysis_Current_Inven', attr_text='Analysis!$S:$S')
analysis_inven_value = range('Analysis_Inven_Value', attr_text='Analysis!$T:$T')
analysis_avg_inven = range('Analysis_Avg_Inven', attr_text='Analysis!$U:$U')
analysis_avg_value = range('Analysis_Avg_Value', attr_text='Analysis!$V:$V')
analysis_annual_volume = range('Analysis_Annual_Volume', attr_text='Analysis!$W:$W')
analysis_demand = range('Analysis_Demand', attr_text='Analysis!$X:$X')
analysis_annual_margin = range('Analysis_Annual_Margin', attr_text='Analysis!$Y:$Y')
analysis_months_active = range('Analysis_Months_Active', attr_text='Analysis!$Z:$Z')
analysis_demand_variability = range('Analysis_Demand_Variability', attr_text='Analysis!$AA:$AA')
analysis_forecast_accuracy = range('Analysis_Forecast_Accuracy', attr_text='Analysis!$AB:$AB')
analysis_ss_demand_variance = range('Analysis_SS_Demand_Variance', attr_text='Analysis!$AC:$AC')
analysis_ss_forecast_error = range('Analysis_SS_Forecast_Error', attr_text='Analysis!$AD:$AD')
analysis_suggested_ss = range('Analysis_Suggested_Ss', attr_text='Analysis!$AE:$AE')
analysis_cycle_stock = range('Analysis_Cycle_Stock', attr_text='Analysis!$AF:$AF')
analysis_min_inventory = range('Analysis_Min_Inventory', attr_text='Analysis!$AG:$AG')
analysis_max_inventory = range('Analysis_Max_Inventory', attr_text='Analysis!$AH:$AH')
analysis_avg_inventory = range('Analysis_Avg_Inventory', attr_text='Analysis!$AI:$AI')
analysis_avg_inventory_value = range('Analysis_Avg_Inventory_Value', attr_text='Analysis!$AJ:$AJ')
analysis_diff = range('Analysis_Diff', attr_text='Analysis!$AK:$AK')

create(analysis_concat)
create(analysis_material)
create(analysis_location)
create(analysis_mat_descr)
create(analysis_location_descr)
create(analysis_include)
create(analysis_purchased)
create(analysis_category)
create(analysis_subcategory)
create(analysis_abc_class)
create(analysis_inven_turns)
create(analysis_avg_days)
create(analysis_velocity_category)
create(analysis_stock_in_months)
create(analysis_stock_category)
create(analysis_abc_class_cat)
create(analysis_unit_price)
create(analysis_unit_cost)
create(analysis_current_inven)
create(analysis_inven_value)
create(analysis_avg_inven)
create(analysis_avg_value)
create(analysis_annual_volume)
create(analysis_demand)
create(analysis_annual_margin)
create(analysis_months_active)
create(analysis_demand_variability)
create(analysis_forecast_accuracy)
create(analysis_ss_demand_variance)
create(analysis_ss_forecast_error)
create(analysis_suggested_ss)
create(analysis_cycle_stock)
create(analysis_min_inventory)
create(analysis_max_inventory)
create(analysis_avg_inventory)
create(analysis_avg_inventory_value)
create(analysis_diff)























# Assign sheets to variables
params = wb["Parameters"]
analysis = wb["Analysis"]
inven_hist = wb["Inventory History"]
cur_inven = wb["Current Inventory"]
sales_hist = wb["Sales History"]
forecast_hist = wb["Forecast History"]
inven_turns = wb["Inventory Turns"]
segmen = wb["Segmentation"]
segmen_calc = wb["Segmentation Calculation"]
material = wb["Material Data"]
skus = wb["SKU Data"]
locale = wb["Location Data"]
conversion = wb["Conversion"]

# Format Analysis
analysis.freeze_panes = "B3"
analysis['A2'] = 'Concat'
analysis['B2'] = 'Material'
analysis['C2'] = 'Location'
analysis['D2'] = 'Material Description'
analysis['E2'] = 'Location Desctiption'
analysis['F2'] = 'Include?'
analysis['G2'] = 'Purchased/Produced?'
analysis['H2'] = 'Category'
analysis['I2'] = 'Sub Category'
analysis['J2'] = 'Abc Class'
analysis['K2'] = 'Inven Turns'
analysis['L2'] = 'Avg Days To Turn Inven'
analysis['M2'] = 'Velocity Category'
analysis['N2'] = 'Stock In Months'
analysis['O2'] = 'Stock Category'
analysis['P2'] = 'Abc Class Category'
analysis['Q2'] = 'Unit Price'
analysis['R2'] = 'Unit Cost'
analysis['S2'] = 'Current Inventory'
analysis['T2'] = 'Inventory Value'
analysis['U2'] = 'Average Inventory'
analysis['V2'] = 'Average Value'
analysis['W2'] = 'Annual Volume'
analysis['X2'] = 'Demand $'
analysis['Y2'] = 'Annual Margin'
analysis['Z2'] = 'Months Active'
analysis['AA2'] = 'Demand Variability'
analysis['AB2'] = 'Fcst Accuracy'
analysis['AC2'] = 'SS - Demand Variance'
analysis['AD2'] = 'SS - Forecast Error'
analysis['AE2'] = 'Suggested SS Method'
analysis['AF2'] = 'Cycle Stock'
analysis['AG2'] = 'Min Inventory'
analysis['AH2'] = 'Max Inventory'
analysis['AI2'] = 'Avg Inventory'
analysis['AJ2'] = 'Avg Inventory Value'
analysis['AK2'] = 'Diff vs Current'

def analysis_header():
    worksheet_names = wb.sheetnames
    sheet_index = worksheet_names.index("Analysis")
    wb.active = sheet_index
    sheet = wb.active
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.align = Alignment(horizontal="center", vertical="center")
    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

# Format Inventory History
inven_hist.freeze_panes = "B3"
inven_hist['A2'] = 'Concat'
inven_hist['B2'] = 'Material'
inven_hist['C2'] = 'Location'
inven_hist['D2'] = 'Month 01'
inven_hist['E2'] = 'Month 02'
inven_hist['F2'] = 'Month 03'
inven_hist['G2'] = 'Month 04'
inven_hist['H2'] = 'Month 05'
inven_hist['I2'] = 'Month 06'
inven_hist['J2'] = 'Month 07'
inven_hist['K2'] = 'Month 08'
inven_hist['L2'] = 'Month 09'
inven_hist['M2'] = 'Month 10'
inven_hist['N2'] = 'Month 11'
inven_hist['O2'] = 'Month 12'
inven_hist['P2'] = 'Month 13'
inven_hist['Q2'] = 'Month 14'
inven_hist['R2'] = 'Average'
inven_hist['S2'] = 'Min'
inven_hist['T2'] = 'Max'
inven_hist['U2'] = 'Total'
inven_hist['V2'] = 'Months with inventory'
inven_hist['W2'] = 'Avg Inventory when >0'
inven_hist['X2'] = 'Min when >0'

def inven_hist_header():
    worksheet_names = wb.sheetnames
    sheet_index = worksheet_names.index("Inventory History")
    wb.active = sheet_index
    sheet = wb.active
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.align = Alignment(horizontal="center", vertical="center")
    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

# Format Current Inventory
cur_inven.freeze_panes = "B3"
cur_inven['A2'] = 'Concat'
cur_inven['B2'] = 'Material'
cur_inven['C2'] = 'Location'
cur_inven['D2'] = 'Current Inventory'
cur_inven['E2'] = 'Inventory value'
cur_inven['F2'] = 'COGS'
cur_inven['G2'] = 'Price'

# Format Sales History
sales_hist.freeze_panes = "B3"
sales_hist['A2'] = 'Concat'
sales_hist['B2'] = 'Material'
sales_hist['C2'] = 'Location'
sales_hist['D2'] = 'Month 01'
sales_hist['E2'] = 'Month 02'
sales_hist['F2'] = 'Month 03'
sales_hist['G2'] = 'Month 04'
sales_hist['H2'] = 'Month 05'
sales_hist['I2'] = 'Month 06'
sales_hist['J2'] = 'Month 07'
sales_hist['K2'] = 'Month 08'
sales_hist['L2'] = 'Month 09'
sales_hist['M2'] = 'Month 10'
sales_hist['N2'] = 'Month 11'
sales_hist['O2'] = 'Month 12'
sales_hist['P2'] = 'Total Demand'
sales_hist['Q2'] = 'Active Months'
sales_hist['R2'] = 'Avg Demand'
sales_hist['S2'] = 'Min'
sales_hist['T2'] = 'Max'
sales_hist['U2'] = 'AVG >0'
sales_hist['V2'] = 'MIN >0'
sales_hist['W2'] = 'Lt'
sales_hist['X2'] = 'Annual'
sales_hist['Y2'] = 'Q1'
sales_hist['Z2'] = 'Q2'
sales_hist['AA2'] = 'Q3'
sales_hist['AB2'] = 'Q4'
sales_hist['AC2'] = 'Annual'
sales_hist['AD2'] = 'Q1'
sales_hist['AE2'] = 'Q2'
sales_hist['AF2'] = 'Q3'
sales_hist['AG2'] = 'Q4'
sales_hist['AH2'] = 'Annual'
sales_hist['AI2'] = 'Q1'
sales_hist['AJ2'] = 'Q2'
sales_hist['AK2'] = 'Q3'

# Format Forecast History
forecast_hist.freeze_panes = "B3"
forecast_hist['A2'] = 'Avg Demand'
forecast_hist['A2'] = 'Concat'
forecast_hist['B2'] = 'Material'
forecast_hist['C2'] = 'Location'
forecast_hist['D2'] = 'Month 01'
forecast_hist['E2'] = 'Month 02'
forecast_hist['F2'] = 'Month 03'
forecast_hist['G2'] = 'Month 04'
forecast_hist['H2'] = 'Month 05'
forecast_hist['I2'] = 'Month 06'
forecast_hist['J2'] = 'Month 07'
forecast_hist['K2'] = 'Month 08'
forecast_hist['L2'] = 'Month 09'
forecast_hist['M2'] = 'Month 10'
forecast_hist['N2'] = 'Month 11'
forecast_hist['O2'] = 'Month 12'
forecast_hist['P2'] = 'Month 01'
forecast_hist['Q2'] = 'Month 02'
forecast_hist['R2'] = 'Month 03'
forecast_hist['S2'] = 'Month 04'
forecast_hist['T2'] = 'Month 05'
forecast_hist['U2'] = 'Month 06'
forecast_hist['V2'] = 'Month 07'
forecast_hist['W2'] = 'Month 08'
forecast_hist['X2'] = 'Month 09'
forecast_hist['Y2'] = 'Month 10'
forecast_hist['Z2'] = 'Month 11'
forecast_hist['AA2'] = 'Month 12'
forecast_hist['AB2'] = 'Month 01'
forecast_hist['AC2'] = 'Month 02'
forecast_hist['AD2'] = 'Month 03'
forecast_hist['AE2'] = 'Month 04'
forecast_hist['AF2'] = 'Month 05'
forecast_hist['AG2'] = 'Month 06'
forecast_hist['AH2'] = 'Month 07'
forecast_hist['AI2'] = 'Month 08'
forecast_hist['AJ2'] = 'Month 09'
forecast_hist['AK2'] = 'Month 10'
forecast_hist['AL2'] = 'Month 11'
forecast_hist['AM2'] = 'Month 12'
forecast_hist['AN2'] = 'Total Bias'
forecast_hist['AO2'] = 'Month 01'
forecast_hist['AP2'] = 'Month 02'
forecast_hist['AQ2'] = 'Month 03'
forecast_hist['AR2'] = 'Month 04'
forecast_hist['AS2'] = 'Month 05'
forecast_hist['AT2'] = 'Month 06'
forecast_hist['AU2'] = 'Month 07'
forecast_hist['AV2'] = 'Month 08'
forecast_hist['AW2'] = 'Month 09'
forecast_hist['AX2'] = 'Month 10'
forecast_hist['AY2'] = 'Month 11'
forecast_hist['AZ2'] = 'Month 12'
forecast_hist['BA2'] = 'RMSE'
forecast_hist['BB2'] = 'RMSE ACTIVE'
forecast_hist['BC2'] = 'Month 01'
forecast_hist['BD2'] = 'Month 02'
forecast_hist['BE2'] = 'Month 03'
forecast_hist['BF2'] = 'Month 04'
forecast_hist['BG2'] = 'Month 05'
forecast_hist['BH2'] = 'Month 06'
forecast_hist['BI2'] = 'Month 07'
forecast_hist['BJ2'] = 'Month 08'
forecast_hist['BK2'] = 'Month 09'
forecast_hist['BL2'] = 'Month 10'
forecast_hist['BM2'] = 'Month 11'
forecast_hist['BN2'] = 'Month 12'
forecast_hist['BO2'] = 'SUM OF ERROR %'
forecast_hist['BP2'] = 'AVG MAPE'
forecast_hist['BQ2'] = 'Avg Bias'
forecast_hist['BR2'] = 'AVG MAPE'
forecast_hist['BS2'] = 'Avg Bias'
forecast_hist['BT2'] = 'MAD'
forecast_hist['BU2'] = 'Tracking Signal'
forecast_hist['BV2'] = 'TS'
forecast_hist['BW2'] = 'TY Accuracy'
forecast_hist['BX2'] = 'Active Accuracy'
forecast_hist['BY2'] = 'SS ANNUAL ERROR'
forecast_hist['BZ2'] = 'SS ACTIVE ERROR'

#Format Inventory Turns
inven_turns.freeze_panes = "B3"
inven_turns['A2'] = 'Concat'
inven_turns['B2'] = 'Material'
inven_turns['C2'] = 'Location'
inven_turns['D2'] = 'Qty On Hand'
inven_turns['E2'] = 'On Hand Value'
inven_turns['F2'] = 'COGS'
inven_turns['G2'] = 'Avg Inven $'
inven_turns['H2'] = 'Inven Turns'
inven_turns['I2'] = 'Days to Turn Inven'
inven_turns['J2'] = 'Velocity Category'
inven_turns['K2'] = 'On Hand Inventory'
inven_turns['L2'] = 'Stock in Months'
inven_turns['M2'] = 'Stock Category'

# Segmentation
segmen.freeze_panes = "B3"
segmen['A2'] = 'Concat'
segmen['B2'] = 'Sku'
segmen['C2'] = 'Loc'
segmen['D2'] = 'Segmentation'
segmen['E2'] = 'Category Segmentation'

# Segmentation Calculation
segmen_calc.freeze_panes = "B3"
segmen_calc['A2'] = 'Concat'
segmen_calc['B2'] = 'Material'
segmen_calc['C2'] = 'Location'
segmen_calc['D2'] = 'Category'
segmen_calc['E2'] = 'Sub Category'
segmen_calc['F2'] = 'Annual Volume'
segmen_calc['G2'] = 'Demand $'
segmen_calc['H2'] = 'Annual Margin'
segmen_calc['I2'] = 'Sku %'
segmen_calc['J2'] = 'Cumulative'

# Material Data
material.freeze_panes = "B3"
material['A2'] = 'Material'
material['B2'] = 'Description'
material['C2'] = 'Unit Cost'
material['D2'] = 'Currency'
material['E2'] = 'Unit Price'
material['F2'] = 'Category'
material['G2'] = 'Subcategory'
material['H2'] = 'Status'
material['I2'] = 'Unit Price'
material['J2'] = 'Units/Case'
material['K2'] = 'Case/Pallet'

# SKU Data
skus.freeze_panes = "B3"
skus['A2'] = 'ConCat'
skus['B2'] = 'Material'
skus['C2'] = 'Locale'
skus['D2'] = 'Purchasing/Producing?'

# Locales
locale.freeze_panes = "B3"
locale['A2'] = 'Location'
locale['B2'] = 'Description'

index_params = 0
index_analysis = 1
index_inven_hist = 2
index_cur_inven = 3
index_sales_hist = 4
index_forecast_hist = 5
index_inven_turns = 6
index_segmen = 7
index_segmen_calc = 8
index_material = 9
index_skus = 10
index_locale = 11
index_conversion = 12

wb.save(filename=filename)